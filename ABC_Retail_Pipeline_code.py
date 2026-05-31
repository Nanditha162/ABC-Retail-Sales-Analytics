
# # ABC Retail Solutions — End-to-End Data Pipeline
# ### Ingestion → Cleaning → Transformation → Curated Dataset for Power BI


# ## 1. Environment Setup
# !pip install pandas openpyxl hashlib

import pandas as pd
import hashlib
import warnings
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')
pd.set_option('display.max_columns', None)
pd.set_option('display.float_format', '{:.2f}'.format)

print("✅ All libraries imported successfully")


# ---
# ## 2. Data Ingestion


# ── Load raw datasets ──────────────────────────────────────
FILE_PATH = 'raw_retail_data.xlsx'   

retail1 = pd.read_excel(FILE_PATH, sheet_name='retail_data1')
retail2 = pd.read_excel(FILE_PATH, sheet_name='retail_data2')
product_dim = pd.read_excel(FILE_PATH, sheet_name='product_details')

print(f"retail_data1   : {retail1.shape[0]:,} rows × {retail1.shape[1]} columns")
print(f"retail_data2   : {retail2.shape[0]:,} rows × {retail2.shape[1]} columns")
print(f"product_dim    : {product_dim.shape[0]:,} rows × {product_dim.shape[1]} columns")


# ── Quick preview ──────────────────────────────────────────
print("=== retail_data1 sample ===")
display(retail1.head(3))
print("\n=== retail_data2 sample ===")
display(retail2.head(3))
print("\n=== product_dim ===")
display(product_dim)


# ── Schema / dtype inspection ──────────────────────────────
print("=== retail_data1 dtypes ===")
print(retail1.dtypes)
print("\n=== retail_data2 dtypes ===")
print(retail2.dtypes)


# ---
# ## 3. Data Quality Checks (Pre-Cleaning)
# 
# Assess missing values, duplicate records, and schema issues before any transformation.

# ── Missing value audit ────────────────────────────────────
def missing_audit(df, name):
    missing = df.isnull().sum()
    pct = (missing / len(df) * 100).round(2)
    audit = pd.DataFrame({'Missing Count': missing, 'Missing %': pct})
    audit = audit[audit['Missing Count'] > 0]
    print(f"\n=== {name} — Missing Values ===")
    if audit.empty:
        print("  No missing values found.")
    else:
        display(audit)

missing_audit(retail1, 'retail_data1')
missing_audit(retail2, 'retail_data2')


# ── Duplicate check ───────────────────────────────────────
print(f"retail_data1 full duplicates  : {retail1.duplicated().sum():,}")
print(f"retail_data2 full duplicates  : {retail2.duplicated().sum():,}")

# Transaction ID duplicates within each source
if 'transaction_id' in retail1.columns:
    print(f"retail_data1 duplicate tx IDs : {retail1['transaction_id'].duplicated().sum():,}")
    print(f"retail_data2 duplicate tx IDs : {retail2['transaction_id'].duplicated().sum():,}")


# ── Basic stats for numeric columns ────────────────────────
print("=== retail_data1 numeric summary ===")
display(retail1.describe())


# ---
# ## 4. Combine Source Datasets
# 
# Stack retail_data1 and retail_data2 into a single DataFrame, tagging each record with its source system.
# 


# ── Tag and concatenate ────────────────────────────────────
retail1['source'] = 'retail_data1'
retail2['source'] = 'retail_data2'

raw = pd.concat([retail1, retail2], ignore_index=True)

print(f"Combined dataset: {raw.shape[0]:,} rows × {raw.shape[1]} columns")
print(f"  retail_data1 records : {(raw['source']=='retail_data1').sum():,}")
print(f"  retail_data2 records : {(raw['source']=='retail_data2').sum():,}")


# ---
# ## 5. Data Cleaning & Transformation
# 
# Step-by-step cleaning pipeline applied to the combined dataset.
# 
# ### 5.1 Remove duplicates
# 


# ── Remove full-row duplicates ─────────────────────────────
before = len(raw)
raw.drop_duplicates(inplace=True)
print(f"Rows before: {before:,}  |  After: {len(raw):,}  |  Removed: {before - len(raw):,}")


# ── Remove duplicate transaction IDs (keep first occurrence) ─
before = len(raw)
raw.drop_duplicates(subset='transaction_id', keep='first', inplace=True)
print(f"Rows before: {before:,}  |  After: {len(raw):,}  |  Removed: {before - len(raw):,}")


# ### 5.2 Standardise column names


# ── Lowercase and strip whitespace from column names ──────
raw.columns = raw.columns.str.strip().str.lower().str.replace(' ', '_')
print("Columns:", raw.columns.tolist())


# ### 5.3 Handle missing values


# ── Drop rows missing critical identifier fields ───────────
critical_cols = ['transaction_id', 'customer_id', 'product_id']
before = len(raw)
raw.dropna(subset=critical_cols, inplace=True)
print(f"After dropping missing critical fields: {len(raw):,} rows (removed {before - len(raw):,})")

# ── Fill missing prices from product_dim ──────────────────
if 'price' in raw.columns:
    price_map = product_dim.set_index('product_id')['price'].to_dict()
    mask = raw['price'].isnull()
    raw.loc[mask, 'price'] = raw.loc[mask, 'product_id'].map(price_map)
    print(f"Filled {mask.sum()} missing price values from product_dim")

# ── Fill missing discounts with 0 ─────────────────────────
if 'discount' in raw.columns:
    filled = raw['discount'].isnull().sum()
    raw['discount'].fillna(0, inplace=True)
    print(f"Filled {filled} missing discount values with 0")

# ── Fill missing quantities with 1 ────────────────────────
if 'quantity' in raw.columns:
    filled = raw['quantity'].isnull().sum()
    raw['quantity'].fillna(1, inplace=True)
    print(f"Filled {filled} missing quantity values with 1")


# ### 5.4 Standardise categorical values


# ── Normalise text fields ──────────────────────────────────
str_cols = ['product_name', 'category', 'city', 'payment_method', 'purchase_location', 'payment_status']
for col in str_cols:
    if col in raw.columns:
        raw[col] = raw[col].astype(str).str.strip().str.title()

# purchase_location → lowercase (online/offline convention)
if 'purchase_location' in raw.columns:
    raw['purchase_location'] = raw['purchase_location'].str.lower()

print("✅ String columns standardised")
print("  category values       :", raw['category'].unique().tolist())
print("  purchase_location     :", raw['purchase_location'].unique().tolist())
print("  payment_method values :", raw['payment_method'].unique().tolist())


# ### 5.5 Fix date formats


# ── Parse transaction_date to datetime ────────────────────
raw['transaction_date'] = pd.to_datetime(raw['transaction_date'], errors='coerce', dayfirst=True)

# Drop rows where date could not be parsed
before = len(raw)
raw.dropna(subset=['transaction_date'], inplace=True)
print(f"Invalid dates dropped: {before - len(raw):,}")
print(f"Date range: {raw['transaction_date'].min().date()} → {raw['transaction_date'].max().date()}")

# Derived date columns for Power BI time intelligence
raw['year_month'] = raw['transaction_date'].dt.to_period('M').astype(str)
raw['year']       = raw['transaction_date'].dt.year
raw['month']      = raw['transaction_date'].dt.month
raw['month_name'] = raw['transaction_date'].dt.strftime('%b')
raw['quarter']    = raw['transaction_date'].dt.quarter.apply(lambda q: f"Q{q}")


# ### 5.6 Validate and fix quantities


# ── Remove invalid quantities (zero or negative) ──────────
before = len(raw)
raw = raw[raw['quantity'] > 0]
print(f"Rows removed for invalid quantity: {before - len(raw):,}")
print(f"Quantity range: {raw['quantity'].min()} – {raw['quantity'].max()}")


# ### 5.7 Enrich from product_dim


# ── Join product_dim to enrich product_name and category ──
product_dim_clean = product_dim.rename(columns={
    'product_id': 'product_id',
    'product_name': 'product_name_ref',
    'category': 'category_ref',
    'price': 'standard_price'
})

raw = raw.merge(
    product_dim_clean[['product_id','product_name_ref','category_ref']],
    on='product_id', how='left'
)

# Use reference values where transaction fields are blank/mismatched
raw['product_name'] = raw['product_name'].fillna(raw['product_name_ref'])
raw['category']     = raw['category'].fillna(raw['category_ref'])

raw.drop(columns=['product_name_ref','category_ref'], inplace=True)

# Flag records not in product_dim
raw['valid_product'] = raw['product_id'].isin(product_dim['product_id'])
print(f"Transactions with unmatched product_id: {(~raw['valid_product']).sum():,}")
raw.drop(columns=['valid_product'], inplace=True)


# ---
# ## 6. PII Masking
# 
# Personally Identifiable Information (PII) must be protected before the dataset is used for analytics.
# 
# - **Email address** → fully masked as `****@****.***`
# - **Phone number** → fully masked as `**********`
# 


# ── Complete masking of email and phone ───────────────────
if 'email' in raw.columns:
    raw.rename(columns={'email': 'email_masked'}, inplace=True)

if 'phone' in raw.columns:
    raw.rename(columns={'phone': 'phone_masked'}, inplace=True)

# Apply full masking
if 'email_masked' in raw.columns:
    raw['email_masked'] = '****@****.***'

if 'phone_masked' in raw.columns:
    raw['phone_masked'] = '**********'

print("✅ PII masking applied")
print("  email sample :", raw['email_masked'].iloc[0])
print("  phone sample :", raw['phone_masked'].iloc[0])


# ---
# ## 7. Revenue Calculation
# 
# Derive the `revenue` column:  
# **revenue = price × quantity × (1 − discount)**
# 


# ── Calculate revenue ─────────────────────────────────────
raw['revenue'] = (raw['price'] * raw['quantity'] * (1 - raw['discount'])).round(2)

print(f"Total Revenue : ₹{raw['revenue'].sum():,.0f}")
print(f"Min Revenue   : ₹{raw['revenue'].min():,.0f}")
print(f"Max Revenue   : ₹{raw['revenue'].max():,.0f}")
print(f"Avg Revenue   : ₹{raw['revenue'].mean():,.2f}")


# ---
# ## 8. Assemble Final Curated Dataset
# 
# Select and order columns for the final `curated_transactions` table.
# 


# ── Select final columns ──────────────────────────────────
final_cols = [
    'transaction_id', 'customer_id', 'customer_name',
    'product_id', 'price', 'product_name', 'category',
    'purchase_location', 'city', 'transaction_date',
    'quantity', 'payment_method', 'discount', 'payment_status',
    'source', 'email_masked', 'phone_masked',
    'revenue', 'year_month', 'year', 'month', 'month_name', 'quarter'
]

# Keep only columns that exist
final_cols = [c for c in final_cols if c in raw.columns]
curated = raw[final_cols].copy()

print(f"Final dataset: {curated.shape[0]:,} rows × {curated.shape[1]} columns")
print(f"Columns: {curated.columns.tolist()}")
display(curated.head(3))


# ---
# ## 9. Business KPI Aggregations
# 
# Pre-compute aggregated summary tables to complement the transaction-level data in Power BI.
# 


# ── KPI Summary ───────────────────────────────────────────
total_rev     = curated['revenue'].sum()
total_orders  = len(curated)
avg_order_val = curated['revenue'].mean()
total_units   = curated['quantity'].sum()
avg_discount  = curated['discount'].mean()
discount_val  = (curated['price'] * curated['quantity'] * curated['discount']).sum()
online_rev    = curated[curated['purchase_location']=='online']['revenue'].sum()
offline_rev   = curated[curated['purchase_location']=='offline']['revenue'].sum()
top_prod      = curated.groupby('product_name')['revenue'].sum().idxmax()
top_prod_rev  = curated.groupby('product_name')['revenue'].sum().max()

kpi_summary = pd.DataFrame({
    'KPI': [
        'Total Revenue', 'Total Orders', 'Average Order Value', 'Total Units Sold',
        'Average Discount Rate', 'Total Discount Value',
        'Online Revenue', 'Offline Revenue',
        'Online Revenue %', 'Offline Revenue %',
        'Top Selling Product', 'Top Product Revenue'
    ],
    'Value': [
        total_rev, total_orders, round(avg_order_val,2), int(total_units),
        round(avg_discount,4), round(discount_val,2),
        online_rev, offline_rev,
        round(online_rev/total_rev*100,2), round(offline_rev/total_rev*100,2),
        top_prod, top_prod_rev
    ]
})

display(kpi_summary)


# ── Revenue by Category ───────────────────────────────────
rev_category = curated.groupby('category').agg(
    revenue=('revenue','sum'),
    orders=('transaction_id','count'),
    units_sold=('quantity','sum')
).reset_index().sort_values('revenue', ascending=False)
rev_category['revenue_pct'] = (rev_category['revenue'] / rev_category['revenue'].sum() * 100).round(2)
display(rev_category)


# ── Revenue by City ───────────────────────────────────────
rev_city = curated.groupby('city').agg(
    revenue=('revenue','sum'),
    orders=('transaction_id','count'),
    units_sold=('quantity','sum')
).reset_index().sort_values('revenue', ascending=False)
display(rev_city)


# ── Revenue by Product ────────────────────────────────────
rev_product = curated.groupby(['product_id','product_name','category']).agg(
    revenue=('revenue','sum'),
    orders=('transaction_id','count'),
    units_sold=('quantity','sum')
).reset_index().sort_values('revenue', ascending=False)
display(rev_product)


# ── Revenue by Channel ────────────────────────────────────
rev_channel = curated.groupby('purchase_location').agg(
    revenue=('revenue','sum'),
    orders=('transaction_id','count')
).reset_index()
display(rev_channel)

# ── Revenue by Payment Method ──────────────────────────────
rev_payment = curated.groupby('payment_method').agg(
    revenue=('revenue','sum'),
    orders=('transaction_id','count')
).reset_index().sort_values('revenue', ascending=False)
display(rev_payment)

# ── Monthly Revenue Trend ──────────────────────────────────
monthly_rev = curated.groupby('year_month').agg(
    revenue=('revenue','sum'),
    orders=('transaction_id','count')
).reset_index().sort_values('year_month')
display(monthly_rev.head(10))

# ── City × Category Revenue Matrix ────────────────────────
city_cat = curated.pivot_table(
    values='revenue', index='city',
    columns='category', aggfunc='sum', fill_value=0
).reset_index()
display(city_cat)


# 


# ── Export all tables to Excel ────────────────────────────
OUTPUT_PATH = 'curated_retail_data_final.xlsx'

with pd.ExcelWriter(OUTPUT_PATH, engine='openpyxl') as writer:
    product_dim.to_excel(writer,   sheet_name='product_dim',          index=False)
    curated.to_excel(writer,       sheet_name='curated_transactions',  index=False)
    kpi_summary.to_excel(writer,   sheet_name='KPI_Summary',           index=False)
    rev_category.to_excel(writer,  sheet_name='Revenue_by_Category',   index=False)
    rev_city.to_excel(writer,      sheet_name='Revenue_by_City',        index=False)
    rev_product.to_excel(writer,   sheet_name='Revenue_by_Product',     index=False)
    rev_channel.to_excel(writer,   sheet_name='Revenue_by_Channel',     index=False)
    rev_payment.to_excel(writer,   sheet_name='Revenue_by_Payment',     index=False)
    monthly_rev.to_excel(writer,   sheet_name='Monthly_Revenue',        index=False)
    city_cat.to_excel(writer,      sheet_name='City_Category_Matrix',   index=False)

    # Auto-fit column widths
    from openpyxl.utils import get_column_letter
    wb = writer.book
    for sname in wb.sheetnames:
        ws = wb[sname]
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

print(f"✅ Exported successfully → {OUTPUT_PATH}")
print(f"   Sheets: product_dim, curated_transactions, KPI_Summary,")
print(f"           Revenue_by_Category, Revenue_by_City, Revenue_by_Product,")
print(f"           Revenue_by_Channel, Revenue_by_Payment, Monthly_Revenue,")
print(f"           City_Category_Matrix")


# ---
# ## 11. Final Validation
# 
# Quick sanity checks on the exported file before loading into Power BI.
# 


# ── Reload and validate ───────────────────────────────────
val = pd.read_excel(OUTPUT_PATH, sheet_name='curated_transactions')

print("=== Final Dataset Validation ===")
print(f"  Rows              : {len(val):,}")
print(f"  Columns           : {val.shape[1]}")
print(f"  Null values       : {val.isnull().sum().sum()}")
print(f"  Duplicate tx IDs  : {val['transaction_id'].duplicated().sum()}")
print(f"  Negative qty      : {(val['quantity'] <= 0).sum()}")
print(f"  Total Revenue     : ₹{val['revenue'].sum():,.0f}")
print(f"  Date range        : {val['transaction_date'].min().date()} → {val['transaction_date'].max().date()}")
print(f"  Cities            : {sorted(val['city'].unique().tolist())}")
print(f"  Categories        : {sorted(val['category'].unique().tolist())}")
print(f"  Email masked      : {val['email_masked'].iloc[0]}")
print(f"  Phone masked      : {val['phone_masked'].iloc[0]}")
print("\n✅ Dataset is clean and Power BI ready.")


# ---
# ## 12. Pipeline Summary
# 
# | Step | Action | Result |
# |------|--------|--------|
# | 1 | Load retail_data1, retail_data2, product_dim | 3 sources ingested |
# | 2 | Data quality checks | Nulls and duplicates identified |
# | 3 | Combine sources | Single DataFrame with source tag |
# | 4 | Remove duplicates | Full row + transaction_id dedup |
# | 5 | Handle missing values | Price from product_dim, discount/qty defaults |
# | 6 | Standardise text/dates | Uniform category, city, date format |
# | 7 | Validate quantities | Invalid records dropped |
# | 8 | Enrich from product_dim | Product name and category validated |
# | 9 | PII masking | Email → `****@****.***`, Phone → `**********` |
# | 10 | Revenue calculation | price × quantity × (1 − discount) |
# | 11 | KPI aggregation | 10 summary tables created |
# | 12 | Export to Excel | `curated_retail_data_final.xlsx` |
# 
# 
#
